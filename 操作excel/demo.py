'''

python 操作excel

'''


import openpyxl

# 创建 Excel 对象
wb = openpyxl.Workbook()

# 保存成文件; 会覆盖已有文件
wb.save(r'test.xlsx')

# 加载xlsx文件
wb = load_workbook('test.xlsx')

'''
	操作工作表
'''

# 获取 Excel 打开后默认的工作表
default_ws = wb.active

# 创建一个新工作表
new_ws = wb.create_sheet(title='new ws')

# 修改工作表的名称
new_ws.title = 'modified ws'

# 迭代工作表中所有行
for row in new_ws.iter_rows():
    pass


'''
	操作单元格
'''
# 给 “F5” 单元格赋值
new_ws['F5'] = 'test'
new_ws['F5'].value = 'hello'
# 取出 “F5” 单元格的值
print(new_ws['F5'].value)

# 得到单元格对象
c = openpyxl.cell.Cell(new_ws)
f5 = new_ws['F5']



wb.save(r'test.xlsx')